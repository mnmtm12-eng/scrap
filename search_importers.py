#!/usr/bin/env /tmp/akva_env2/bin/python3
"""
آلة حفر الأرقام - تعمل باستمرار حتى توقفها
Steel door importer hunter - runs forever until stopped
"""

import os, re, csv, json, threading, time, sys, random, argparse
from datetime import date, datetime
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from bs4 import BeautifulSoup
from ddgs import DDGS

DATA_DIR = os.path.dirname(os.path.abspath(__file__))
PID_FILE = os.path.join(DATA_DIR, "runner.pid")
HISTORY_FILE = os.path.join(DATA_DIR, "seen_urls.json")
RESULTS_FILE = os.path.join(DATA_DIR, "results.csv")
EXCEL_FILE = os.path.join(DATA_DIR, "contacts.xlsx")

STEEL_DOOR_KW = [
    "باب مصفح","ابواب مصفحة","باب حديد","ابواب حديدية",
    "steel door","safety door","security door","fire door",
    "باب تركي","ابواب تركية","door importer","door supplier",
    "موزع ابواب","مستورد ابواب","steel security door",
    "باب مقاوم","باب ضد السرقة","باب امان","باب فولاذي",
    "metal door","armored door","hollow metal","fire rated",
]

COUNTRIES = [
    ("الاردن", [
        "site:facebook.com ابواب مصفحة تركيا الاردن",
        "ابواب مصفحة تركية الاردن واتس",
        "وكيل ابواب تركية في الاردن",
        "steel door importer jordan",
        "معرض ابواب مصفحة عمان",
        "شركة ابواب تركية الاردن",
        "site:instagram.com ابواب مصفحة الاردن",
        "باب امان تركي الاردن معارض",
        "site:opensooq.com ابواب مصفحة الاردن",
        "تجار الحديد والابواب في عمان",
    ]),
    ("مصر", [
        "site:facebook.com ابواب مصفحة من تركيا مصر",
        "مستورد ابواب مصفحة من تركيا القاهرة",
        "ابواب مصفحة تركية مصر واتساب",
        "steel door importer egypt",
        "تجار ابواب حديدية في مصر",
        "وكيل باب مصفح تركي في مصر",
        "site:instagram.com ابواب مصفحة مصر",
        "معارض بيع الابواب في مصر",
        "site:opensooq.com ابواب مصفحة مصر",
        "ابواب شقة مصفحة وكلاء مصر",
    ]),
    ("العراق", [
        "site:facebook.com ابواب مصفحة تركيا العراق",
        "ابواب مصفحة تركية للعراق واتس",
        "مستورد ابواب من تركيا لبغداد",
        "شركة ابواب تركية العراق",
        "steel door iraq importer",
        "موزع ابواب مصفحة في بغداد",
        "site:instagram.com ابواب مصفحة العراق",
        "ابواب حديد بغداد",
        "معرض ابواب في العراق",
        "اسعار الابواب المصفحة في العراق",
    ]),
    ("المغرب", [
        "site:facebook.com ابواب مصفحة تركيا المغرب",
        "ابواب مصفحة من تركيا المغرب واتس",
        "مستورد ابواب تركية الدار البيضاء",
        "steel door importer morocco",
        "porte blindee maroc import turquie",
        "شركة استيراد ابواب من تركيا المغرب",
        "site:avito.ma ابواب مصفحة",
        "معرض ابواب في الدار البيضاء",
        "site:instagram.com ابواب مصفحة المغرب",
        "serrurerie porte blindee casablanca",
    ]),
    ("ليبيا", [
        "site:facebook.com ابواب مصفحة تركيا ليبيا",
        "ابواب مصفحة تركية ليبيا واتساب",
        "مستورد ابواب من تركيا طرابلس",
        "steel door importer libya",
        "باب مصفح تركي للبيع في ليبيا",
        "شركة ابواب تركية ليبيا",
        "site:instagram.com ابواب مصفحة ليبيا",
        "موزع ابواب تركية ليبيا",
        "اسعار الابواب المصفحة في ليبيا",
        "تجار مواد بناء في ليبيا",
    ]),
    ("افريقيا", [
        "site:facebook.com steel door supplier africa",
        "steel door importer ghana nigeria kenya",
        "site:alibaba.com steel door africa importer",
        "fire door distributor senegal ivory coast",
        "metal door importer tanzania ethiopia",
        "safety door supplier west africa",
        "site:exportersindia.com steel door africa",
        "armored door supplier south africa kenya",
        "site:facebook.com fire door distributor africa",
        "metal door dealer ghana nigeria cameroon",
    ]),
    ("عام", [
        "site:facebook.com ابواب مصفحة تركيا",
        "مصنع ابواب مصفحة في تركيا واتساب",
        "steel door turkey supplier whatsapp",
        "شركات تصدير ابواب تركية للعالم",
        "site:alibaba.com steel security door turkey",
        "best steel door manufacturer istanbul",
        "wholesale steel door turkey exporter",
        "موردين ابواب مصفحة في تركيا",
        "site:turkishmanufacturers.com door",
        "steel door factory turkey contact",
    ]),
]

VISITED = set()
if os.path.exists(HISTORY_FILE):
    with open(HISTORY_FILE) as f:
        VISITED = set(json.load(f))

CYCLE_RESULTS = []
LOCK = threading.Lock()
HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
SKIP_DOMAINS = ["wikipedia.org","youtube.com","reddit.com","pinterest","amazon","ebay","twitter.com","x.com"]

CYCLE_COUNT = 0
TOTAL_THIS_SESSION = 0


def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


def is_relevant(text):
    t = text.lower()
    return sum(1 for kw in STEEL_DOOR_KW if kw.lower() in t) >= 2


def extract_phones(text):
    phones = set()
    pats = [
        r'(?:\+?\d{1,3}[-.\s]?)?\(?\d{2,4}\)?[-.\s]?\d{3,4}[-.\s]?\d{3,4}',
        r'(?:واتس|اتصل|هاتف|جوال|رقم|موبايل|call|phone|mobile|whatsapp|contact|tel|tel:|phone:)[:\s]*\+?\d{7,15}',
        r'wa\.me/(\+?\d+)',
        r'api\.whatsapp\.com/send\?phone=(\+?\d+)',
    ]
    for pat in pats:
        for m in re.finditer(pat, text, re.IGNORECASE):
            num = re.sub(r'[^\d+]', '', m.group(1) if m.lastindex else m.group())
            if num.startswith("00"): num = "+" + num[2:]
            clean = re.sub(r'[^\d]', '', num)
            if 8 < len(clean) < 16:
                phones.add(clean)
    return phones


def extract_emails(text):
    emails = set()
    for m in re.finditer(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text):
        e = m.group().lower()
        if not any(s in e for s in ["example","domain.com","yoursite","yourdomain","test@"]):
            emails.add(e)
    return emails


def extract_name(soup, url):
    for sel in ["h1", "meta[property='og:title']", "meta[name='twitter:title']"]:
        el = soup.select_one(sel)
        if el and (t := el.get("content", "") or el.get_text(strip=True)):
            if len(t) > 5:
                return t[:80].replace(",", "").replace("|", "-")
    if (title := soup.find("title")):
        parts = [p.strip() for p in re.split(r'[-|–—/\[\]()]', title.get_text(strip=True)) if p.strip()]
        if parts: return parts[0][:80].replace(",", "").replace("|", "-")
    return urlparse(url).netloc.replace("www.", "")


def skip_url(url):
    d = urlparse(url).netloc.lower()
    return any(s in d for s in SKIP_DOMAINS) or any(url.lower().endswith(e) for e in [".pdf",".docx","jpg",".png",".mp4",".gif"])


def search_ddg(query, max_results=10):
    try:
        with DDGS() as ddgs:
            return list(ddgs.text(query, max_results=max_results))
    except:
        return []


def scrape_one(url, country, query):
    if url in VISITED or skip_url(url):
        return
    try:
        resp = requests.get(url, headers=HEADERS, timeout=12)
        if resp.status_code != 200 or "text/html" not in resp.headers.get("content-type",""):
            return
        soup = BeautifulSoup(resp.text, "lxml")
        text = soup.get_text(separator=" ", strip=True)
        if not is_relevant(text):
            return
        phones = extract_phones(text)
        if not phones:
            return
        emails = extract_emails(text)
        name = extract_name(soup, url)
        with LOCK:
            VISITED.add(url)
            CYCLE_RESULTS.append({
                "name": name,
                "country": country,
                "query": query,
                "url": url[:250],
                "phones": ", ".join(sorted(phones))[:500],
                "emails": ", ".join(sorted(emails))[:500],
            })
        log(f"    ✓ {name[:40]:40s} | {len(phones)} رقم | {country}")
    except:
        pass


def save_cycle():
    global CYCLE_RESULTS, TOTAL_THIS_SESSION
    if not CYCLE_RESULTS:
        return
    total_phones = sum(len(r["phones"].split(", ")) for r in CYCLE_RESULTS if r["phones"])
    TOTAL_THIS_SESSION += len(CYCLE_RESULTS)
    keys = ["name","country","query","url","phones","emails"]
    new = not os.path.exists(RESULTS_FILE)
    with open(RESULTS_FILE, "a", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        if new: w.writeheader()
        w.writerows(CYCLE_RESULTS)
    with open(HISTORY_FILE, "w") as f:
        json.dump(list(VISITED), f)
    log(f"  ✓ دورة #{CYCLE_COUNT}: {len(CYCLE_RESULTS)} صفحة | {total_phones} رقم")
    rebuild_excel()
    CYCLE_RESULTS = []


def rebuild_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return

    if not os.path.exists(RESULTS_FILE):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "الأرقام"

    headers = ["#","الاسم","الدولة","أرقام الهاتف","إيميل","الرابط","مصدر البحث"]
    hf = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    for c,h in enumerate(headers,1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hf; cell.font = hfont

    seen = set()
    rn = 2
    with open(RESULTS_FILE, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            pk = re.sub(r'[^\d]','', r.get("phones",""))
            if pk in seen: continue
            seen.add(pk)
            ws.cell(rn,1,rn-1)
            ws.cell(rn,2,r.get("name",""))
            ws.cell(rn,3,r.get("country",""))
            ws.cell(rn,4,r.get("phones",""))
            ws.cell(rn,5,r.get("emails",""))
            ws.cell(rn,6,r.get("url","")[:80])
            ws.cell(rn,7,r.get("query","")[:50])
            for col in range(1,8):
                ws.cell(rn,col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            rn += 1

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{rn-1}"

    wb.save(EXCEL_FILE)
    log(f"  ✓ اكسل: {rn-2} جهة اتصال فريدة")


def run_cycle():
    global CYCLE_COUNT
    CYCLE_COUNT += 1
    log(f"{'='*55}")
    log(f"  دورة #{CYCLE_COUNT} | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    log(f"{'='*55}")
    random.shuffle(COUNTRIES)
    for country, queries in COUNTRIES:
        random.shuffle(queries)
        log(f"\n── {country} ──")
        for q in queries[:6]:
            log(f"  🔍 {q}")
            results = search_ddg(q, max_results=8)
            if results:
                with ThreadPoolExecutor(max_workers=4) as pool:
                    futures = [pool.submit(scrape_one, r.get("href",""), country, q) for r in results if r.get("href","") and r["href"] not in VISITED]
                    for f in as_completed(futures):
                        f.result()
            time.sleep(random.uniform(0.3, 0.7))
    save_cycle()


def main():
    global TOTAL_THIS_SESSION
    parser = argparse.ArgumentParser()
    parser.add_argument("--oneshot", action="store_true", help="تشغيل دورة واحدة ثم الخروج (لـ GitHub Actions)")
    parser.add_argument("--country", help="دولة محددة فقط (مثال: --country مصر)")
    args = parser.parse_args()

    if args.oneshot:
        log("⚡ وضع الدورة الواحدة (CI)")
        random.shuffle(COUNTRIES)
        for country, queries in COUNTRIES:
            if args.country and country != args.country:
                continue
            random.shuffle(queries)
            log(f"\n── {country} ──")
            for q in queries[:3]:
                log(f"  🔍 {q}")
                results = search_ddg(q, max_results=5)
                if results:
                    with ThreadPoolExecutor(max_workers=3) as pool:
                        futures = [pool.submit(scrape_one, r.get("href",""), country, q) for r in results if r.get("href","") and r["href"] not in VISITED]
                        for f in as_completed(futures):
                            f.result()
                time.sleep(random.uniform(0.3, 0.5))
        save_cycle()
        log("✅ انتهت دورة CI")
        return

    with open(PID_FILE, "w") as f:
        f.write(str(os.getpid()))
    log(f"🚀 بدأت آلة الحفر | PID: {os.getpid()}")
    log(f"   {len(VISITED)} موقع مفحوص سابقاً")
    log(f"   لأيقافها: kill {os.getpid()}  أو 双击 stop.sh")
    try:
        while True:
            run_cycle()
            wait = random.randint(300, 900)
            log(f"\n💤 استراحة {wait//60} دقيقة... (Ctrl+C لإيقاف)")
            time.sleep(wait)
    except KeyboardInterrupt:
        log("\n🛑 أوقفها المستخدم")
    finally:
        if os.path.exists(PID_FILE):
            os.remove(PID_FILE)
        log(f"📊 إجمالي هذه الجلسة: {TOTAL_THIS_SESSION} نتيجة")
        log(f"📊 إجمالي المواقع المفحوصة: {len(VISITED)}")
        log(f"👋 السلام!")

if __name__ == "__main__":
    main()
